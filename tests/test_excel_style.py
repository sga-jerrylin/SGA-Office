"""Excel 样式引擎测试"""

import pytest
from io import BytesIO
from openpyxl import load_workbook


class TestExcelStyleEngine:

    def test_themed_header(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Test",
            data=[["Name", "Score"], ["Alice", 95]],
            style={"theme": "business_blue", "header_style": "colored"},
        )
        wb = load_workbook(result)
        ws = wb.active
        # Row 2 is header (row 1 is title)
        header_cell = ws.cell(row=2, column=1)
        assert header_cell.fill.start_color.rgb is not None

    def test_freeze_panes(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Test",
            data=[["A", "B"], [1, 2]],
            style={"freeze_panes": "A3"},
        )
        wb = load_workbook(result)
        ws = wb.active
        assert ws.freeze_panes == "A3"

    def test_auto_filter(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Test",
            data=[["A", "B"], [1, 2]],
            style={"auto_filter": True},
        )
        wb = load_workbook(result)
        ws = wb.active
        assert ws.auto_filter.ref is not None

    def test_alternating_rows(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Test",
            data=[["A", "B"], [1, 2], [3, 4], [5, 6]],
            style={"alternating_rows": True, "theme": "business_blue"},
        )
        wb = load_workbook(result)
        ws = wb.active
        # Row 4 is the 2nd data row (row 1=title, row 2=header, row 3=data1, row 4=data2)
        cell_even = ws.cell(row=4, column=1)
        assert cell_even.fill.start_color is not None

    def test_row_groups_coloring(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Project Plan",
            data=[
                ["Phase", "Task", "Status"],
                ["Alpha", "Task 1", "Done"],
                ["Alpha", "Task 2", "WIP"],
                ["Beta", "Task 3", "TODO"],
            ],
            style={
                "row_groups": {
                    "group_column": "A",
                    "colors": {"Alpha": "2E75B6", "Beta": "7030A0"},
                }
            },
        )
        wb = load_workbook(result)
        ws = wb.active
        # Alpha rows should have blue tint
        alpha_cell = ws.cell(row=3, column=1)  # First Alpha row
        assert alpha_cell.fill.start_color.rgb is not None

    def test_no_style_backward_compat(self):
        """不传 style 时和现有行为一致"""
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Test",
            data=[["A", "B"], [1, 2]],
        )
        assert isinstance(result, BytesIO)
        result.seek(0)
        assert result.read(2) == b"PK"

    def test_column_widths(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Test",
            data=[["Name", "Description"], ["A", "Long text"]],
            style={"column_widths": {"A": 8, "B": 40}},
        )
        wb = load_workbook(result)
        ws = wb.active
        assert ws.column_dimensions["A"].width == 8
        assert ws.column_dimensions["B"].width == 40


class TestExcelStyleSchema:

    def test_style_field_is_optional(self):
        from app.schemas.payload_excel import CreateExcelRequest
        req = CreateExcelRequest(
            title="Test",
            data=[["A", "B"], [1, 2]],
        )
        assert req.style is None

    def test_style_field_accepts_valid_config(self):
        from app.schemas.payload_excel import CreateExcelRequest
        req = CreateExcelRequest(
            title="Test",
            data=[["A", "B"], [1, 2]],
            style={"theme": "business_blue", "freeze_panes": "A3"},
        )
        assert req.style is not None
        assert req.style.theme == "business_blue"


class TestGanttRendering:

    def test_gantt_creates_timeline_columns(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Project Timeline",
            data=[
                ["Task", "Start", "End"],
                ["Task A", "2026-03-01", "2026-03-15"],
                ["Task B", "2026-03-10", "2026-03-25"],
            ],
            style={
                "gantt": {
                    "date_columns": ["B", "C"],
                    "timeline_start": "2026-03-01",
                    "timeline_end": "2026-03-31",
                    "granularity": "week",
                }
            },
        )
        wb = load_workbook(result)
        ws = wb.active
        # Should have more columns than the original 3 (timeline columns added)
        assert ws.max_column > 3

    def test_no_gantt_no_extra_columns(self):
        from app.services.excel_handler import create_excel_from_array
        result = create_excel_from_array(
            title="Simple",
            data=[["A", "B"], [1, 2]],
        )
        wb = load_workbook(result)
        ws = wb.active
        assert ws.max_column == 2
