"""
Excel 核心业务处理引擎
将原先 main.py 中的 Excel 生成逻辑提取、增强，并实现全部 EXC-01~04 功能。
所有方法均为纯函数/IO函数，不涉及 HTTP 层逻辑。
"""

import re
import logging
from io import BytesIO
from datetime import datetime, date, timedelta
from typing import Any, Optional, Union

import requests
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

logger = logging.getLogger(__name__)


# =====================================================
#  通用样式工具
# =====================================================

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_SUMMARY_FILL = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

_SUMMARY_KEYWORDS = {"合计", "总计", "Total", "小计", "Subtotal"}


def _apply_cell_style(
    cell,
    is_header: bool = False,
    is_summary: bool = False,
) -> None:
    """为单元格应用统一的对齐、边框与条件样式"""
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER
    if is_header:
        cell.font = Font(bold=True, size=11)
        cell.fill = _HEADER_FILL
    if is_summary:
        cell.font = Font(bold=True, size=11)
        cell.fill = _SUMMARY_FILL


def _calculate_title_style(title: str) -> dict:
    """根据标题长度动态计算字号与行高"""
    length = len(title)
    if length <= 15:
        return {"size": 16, "height": 25}
    elif length <= 25:
        return {"size": 14, "height": 35}
    elif length <= 40:
        return {"size": 12, "height": 45}
    else:
        return {"size": 11, "height": 55}


def _auto_column_widths(sheet) -> None:
    """自动根据内容设置列宽"""
    for col_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    # 中文字符按 2 字宽计算
                    val = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, char_len)
            except Exception:
                pass
        adjusted = min(max(max_len + 3, 8), 60)  # 最小 8, 最大 60
        sheet.column_dimensions[col_letter].width = adjusted


def _sanitize_filename(name: str, max_length: int = 30) -> str:
    """清理文件名：移除非法字符，截断过长名称"""
    clean = re.sub(r'[\\/:*?"<>|\s]', "", name)[:max_length]
    return clean or "未命名"


# =====================================================
#  Style Engine
# =====================================================

def _parse_excel_style(style: Any) -> Any:
    """Convert a dict to ExcelStyle if needed, or return None."""
    if style is None:
        return None
    # If already an ExcelStyle instance, return as-is
    from app.schemas.payload_excel import ExcelStyle
    if isinstance(style, ExcelStyle):
        return style
    if isinstance(style, dict):
        return ExcelStyle(**style)
    return None


def _apply_style_engine(
    sheet,
    data: list[list[Any]],
    style_config,
    title_row: bool = True,
) -> None:
    """
    Apply the style engine to a sheet.

    Args:
        sheet: openpyxl worksheet
        data: the original data (data[0] = headers, data[1:] = rows)
        style_config: ExcelStyle instance
        title_row: whether a title row exists at row 1 (EXC-01 has one, EXC-03 does not)
    """
    from app.core.themes import get_theme

    theme = get_theme(style_config.theme)
    headers = data[0]
    data_rows = data[1:]
    num_cols = len(headers)
    header_row_idx = 2 if title_row else 1
    first_data_row = header_row_idx + 1

    # --- a) Themed header colors ---
    if style_config.header_style == "colored":
        header_fill = PatternFill(
            start_color=theme.table_header_bg,
            end_color=theme.table_header_bg,
            fill_type="solid",
        )
        header_font = Font(
            bold=True, size=11,
            color=theme.table_header_font,
        )
        for col_num in range(1, num_cols + 1):
            cell = sheet.cell(row=header_row_idx, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
    elif style_config.header_style == "minimal":
        for col_num in range(1, num_cols + 1):
            cell = sheet.cell(row=header_row_idx, column=col_num)
            cell.font = Font(bold=True, size=11)
    # bold_only: just bold, no fill changes (default _apply_cell_style already does bold)

    # --- b) Freeze panes ---
    if style_config.freeze_panes:
        sheet.freeze_panes = style_config.freeze_panes

    # --- c) Auto filter ---
    if style_config.auto_filter:
        end_col_letter = get_column_letter(num_cols)
        sheet.auto_filter.ref = f"A{header_row_idx}:{end_col_letter}{header_row_idx}"

    # --- d) Column widths ---
    if style_config.column_widths:
        for col_letter, width in style_config.column_widths.items():
            sheet.column_dimensions[col_letter].width = width

    # --- e) Alternating rows ---
    if style_config.alternating_rows and style_config.theme:
        alt_fill = PatternFill(
            start_color=theme.table_alt_row_bg,
            end_color=theme.table_alt_row_bg,
            fill_type="solid",
        )
        for i, _ in enumerate(data_rows):
            row_idx = first_data_row + i
            if i % 2 == 1:  # every other row (0-indexed: 1, 3, 5...)
                for col_num in range(1, num_cols + 1):
                    sheet.cell(row=row_idx, column=col_num).fill = alt_fill

    # --- f) Row groups ---
    if style_config.row_groups:
        rg = style_config.row_groups
        # If rg is a dict (shouldn't happen after _parse but be safe)
        from app.schemas.payload_excel import RowGroupConfig
        if isinstance(rg, dict):
            rg = RowGroupConfig(**rg)
        group_col_idx = column_index_from_string(rg.group_column)
        colors_map = rg.colors or {}
        # Auto-assign colors for unknown groups
        auto_colors = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47"]
        auto_idx = 0
        for i, row_data in enumerate(data_rows):
            row_idx = first_data_row + i
            if group_col_idx - 1 < len(row_data):
                group_val = str(row_data[group_col_idx - 1]) if row_data[group_col_idx - 1] is not None else ""
                if group_val and group_val not in colors_map:
                    colors_map[group_val] = auto_colors[auto_idx % len(auto_colors)]
                    auto_idx += 1
                if group_val in colors_map:
                    # Apply a light tint by using the color directly
                    color_hex = colors_map[group_val]
                    # Create a lighter version for row background
                    light_hex = _lighten_color(color_hex, factor=0.7)
                    row_fill = PatternFill(
                        start_color=light_hex,
                        end_color=light_hex,
                        fill_type="solid",
                    )
                    for col_num in range(1, num_cols + 1):
                        sheet.cell(row=row_idx, column=col_num).fill = row_fill

    # --- g) Gantt timeline ---
    if style_config.gantt:
        gantt = style_config.gantt
        from app.schemas.payload_excel import GanttConfig
        if isinstance(gantt, dict):
            gantt = GanttConfig(**gantt)
        _render_gantt_timeline(sheet, data, headers, gantt, style_config, title_row)


def _lighten_color(hex_color: str, factor: float = 0.7) -> str:
    """Lighten a hex color by mixing with white."""
    hex_color = hex_color.lstrip("#")
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def _render_gantt_timeline(
    sheet,
    data: list[list[Any]],
    headers: list,
    gantt_config,
    style_config,
    title_row: bool = True,
) -> None:
    """
    Render a Gantt timeline as colored cells to the right of the data columns.

    Args:
        sheet: openpyxl worksheet
        data: original data (data[0]=headers, data[1:]=rows)
        headers: header row
        gantt_config: GanttConfig instance
        style_config: ExcelStyle instance (for theme colors)
        title_row: whether row 1 is a title row
    """
    from app.core.themes import get_theme

    theme = get_theme(style_config.theme)
    data_rows = data[1:]
    num_data_cols = len(headers)
    header_row_idx = 2 if title_row else 1
    first_data_row = header_row_idx + 1

    # Parse timeline boundaries
    tl_start = datetime.strptime(gantt_config.timeline_start, "%Y-%m-%d").date()
    tl_end = datetime.strptime(gantt_config.timeline_end, "%Y-%m-%d").date()

    # Calculate time slots
    slots = _calculate_time_slots(tl_start, tl_end, gantt_config.granularity)
    if not slots:
        return

    # Write timeline header columns
    start_col = num_data_cols + 1
    for slot_idx, (slot_start, slot_end, label) in enumerate(slots):
        col = start_col + slot_idx
        cell = sheet.cell(row=header_row_idx, column=col, value=label)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        sheet.column_dimensions[get_column_letter(col)].width = 6

    # Parse date columns
    start_col_idx = column_index_from_string(gantt_config.date_columns[0])
    end_col_idx = column_index_from_string(gantt_config.date_columns[1])

    # Default bar color
    default_fill = PatternFill(
        start_color=theme.table_header_bg,
        end_color=theme.table_header_bg,
        fill_type="solid",
    )

    # Fill gantt bars for each data row
    for i, row_data in enumerate(data_rows):
        row_idx = first_data_row + i
        # Get start/end dates from data
        try:
            raw_start = row_data[start_col_idx - 1]
            raw_end = row_data[end_col_idx - 1]
            if raw_start is None or raw_end is None:
                continue
            task_start = _parse_date_value(raw_start)
            task_end = _parse_date_value(raw_end)
            if task_start is None or task_end is None:
                continue
        except (IndexError, ValueError):
            continue

        # Determine bar color
        bar_fill = default_fill
        if gantt_config.bar_color_column and style_config.row_groups:
            # Try to get color from row_groups mapping
            rg = style_config.row_groups
            from app.schemas.payload_excel import RowGroupConfig
            if isinstance(rg, dict):
                rg = RowGroupConfig(**rg)
            bc_idx = column_index_from_string(gantt_config.bar_color_column)
            if bc_idx - 1 < len(row_data):
                group_val = str(row_data[bc_idx - 1]) if row_data[bc_idx - 1] is not None else ""
                if rg.colors and group_val in rg.colors:
                    color = rg.colors[group_val]
                    bar_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Fill cells in the timeline that overlap with task
        for slot_idx, (slot_start, slot_end, _label) in enumerate(slots):
            col = num_data_cols + 1 + slot_idx
            if task_start <= slot_end and task_end >= slot_start:
                sheet.cell(row=row_idx, column=col).fill = bar_fill
                sheet.cell(row=row_idx, column=col).border = _THIN_BORDER


def _parse_date_value(val) -> Optional[date]:
    """Parse a date value from various formats."""
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _calculate_time_slots(
    start: date, end: date, granularity: str
) -> list[tuple[date, date, str]]:
    """
    Calculate time slots for the gantt timeline.
    Returns list of (slot_start, slot_end, label) tuples.
    """
    slots = []
    current = start

    if granularity == "day":
        while current <= end:
            slots.append((current, current, current.strftime("%m/%d")))
            current += timedelta(days=1)

    elif granularity == "week":
        # Align to week start (Monday)
        week_start = current - timedelta(days=current.weekday())
        if week_start < start:
            week_start = start
        while week_start <= end:
            week_end = week_start + timedelta(days=6)
            if week_end > end:
                week_end = end
            label = week_start.strftime("%m/%d")
            slots.append((week_start, week_end, label))
            week_start = week_start + timedelta(days=7)
            # Re-align if we started mid-week
            if len(slots) == 1 and current.weekday() != 0:
                week_start = current - timedelta(days=current.weekday()) + timedelta(days=7)

    elif granularity == "month":
        while current <= end:
            # Month start
            month_start = current.replace(day=1)
            if month_start < start:
                month_start = start
            # Month end
            if current.month == 12:
                month_end = current.replace(year=current.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = current.replace(month=current.month + 1, day=1) - timedelta(days=1)
            if month_end > end:
                month_end = end
            label = current.strftime("%Y-%m")
            slots.append((month_start, month_end, label))
            # Move to next month
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1, day=1)
            else:
                current = current.replace(month=current.month + 1, day=1)
    else:
        # Default to week
        return _calculate_time_slots(start, end, "week")

    return slots


# =====================================================
#  EXC-01: create_excel_from_array
# =====================================================

def create_excel_from_array(
    title: str,
    data: list[list[Any]],
    sheet_name: str = "Sheet1",
    style: Any = None,
) -> BytesIO:
    """
    从二维数组创建简单 Excel 文件。

    Args:
        title:      表格标题（首行合并居中）
        data:       二维数组，data[0] 为表头行
        sheet_name: Sheet 页签名称
        style:      可选的样式配置（dict 或 ExcelStyle 实例）

    Returns:
        BytesIO 对象，包含生成的 .xlsx 数据
    """
    style_config = _parse_excel_style(style)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    headers = data[0]
    data_rows = data[1:]
    end_col = get_column_letter(len(headers))

    # ---------- 标题行 ----------
    title_style = _calculate_title_style(title)
    sheet["A1"] = title
    title_cell = sheet["A1"]
    title_cell.font = Font(size=title_style["size"], bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.merge_cells(f"A1:{end_col}1")
    sheet.row_dimensions[1].height = title_style["height"]

    # ---------- 表头行 ----------
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num, value=header)
        _apply_cell_style(cell, is_header=True)

    # ---------- 数据行 ----------
    for row_idx, row_data in enumerate(data_rows, 3):
        is_summary = bool(row_data and str(row_data[0]) in _SUMMARY_KEYWORDS)
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_num, value=value)
            _apply_cell_style(cell, is_summary=is_summary)

    # ---------- 自动列宽 ----------
    _auto_column_widths(sheet)

    # ---------- Style engine (optional) ----------
    if style_config:
        # Column widths override: skip auto for specified columns
        if style_config.column_widths:
            for col_letter, width in style_config.column_widths.items():
                sheet.column_dimensions[col_letter].width = width
        _apply_style_engine(sheet, data, style_config, title_row=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =====================================================
#  EXC-02: append_rows_to_excel
# =====================================================

def _download_excel_from_url(url: str) -> BytesIO:
    """从公网 URL 下载 Excel 文件到内存"""
    logger.info(f"正在下载 Excel 文件: {url}")
    resp = requests.get(
        str(url),
        timeout=60,
        headers={"User-Agent": "SGA-Office/1.0"},
    )
    resp.raise_for_status()
    return BytesIO(resp.content)


def append_rows_to_excel(
    source_excel_url: str,
    rows: list[list[Any]],
    sheet_name: str = "Sheet1",
) -> BytesIO:
    """
    下载已有 Excel，在指定 Sheet 末尾追加行数据，返回新的 BytesIO。

    Args:
        source_excel_url: 源文件公网 URL
        rows:             要追加的行数据
        sheet_name:       目标 Sheet 名称

    Returns:
        BytesIO 对象，包含追加后的 .xlsx 数据
    """
    excel_data = _download_excel_from_url(source_excel_url)
    wb = openpyxl.load_workbook(excel_data)

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        # 如果指定 Sheet 不存在，使用第一个 Sheet
        logger.warning(f"Sheet '{sheet_name}' 不存在，使用第一个 Sheet: {wb.sheetnames[0]}")
        sheet = wb.active

    # 找到当前最后一行
    max_row = sheet.max_row

    for row_offset, row_data in enumerate(rows, 1):
        target_row = max_row + row_offset
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=target_row, column=col_num, value=value)
            _apply_cell_style(cell)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =====================================================
#  EXC-03: generate_complex_excel
# =====================================================

def generate_complex_excel(
    title: str,
    sheets_def: list[dict],
    style: Any = None,
) -> BytesIO:
    """
    生成包含多个 Sheet、合并单元格、预埋公式的行业级报表。

    Args:
        title:      报表总标题（会写入第一个 Sheet 的文件属性）
        sheets_def: SheetDefinition 列表 (已转 dict)
        style:      全局样式配置（dict 或 ExcelStyle），各 Sheet 可单独覆盖

    Returns:
        BytesIO 对象
    """
    global_style = _parse_excel_style(style)
    wb = openpyxl.Workbook()
    wb.properties.title = title

    for idx, sdef in enumerate(sheets_def):
        if idx == 0:
            sheet = wb.active
            sheet.title = sdef["sheet_name"]
        else:
            sheet = wb.create_sheet(title=sdef["sheet_name"])

        headers = sdef["headers"]
        data_rows = sdef["data"]
        end_col = get_column_letter(len(headers))

        # ---------- 表头行 (Row 1) ----------
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            _apply_cell_style(cell, is_header=True)

        # ---------- 数据行 ----------
        for row_idx, row_data in enumerate(data_rows, 2):
            is_summary = bool(row_data and str(row_data[0]) in _SUMMARY_KEYWORDS)
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_num)

                # 如果值以 '=' 开头，认为是 Excel 公式
                if isinstance(value, str) and value.startswith("="):
                    cell.value = value  # openpyxl 会自动识别公式
                else:
                    cell.value = value

                _apply_cell_style(cell, is_summary=is_summary)

        # ---------- 合并单元格 ----------
        merge_cells = sdef.get("merge_cells") or []
        for merge_range in merge_cells:
            try:
                sheet.merge_cells(merge_range)
            except Exception as e:
                logger.warning(f"合并单元格 '{merge_range}' 失败: {e}")

        # ---------- 自动列宽 ----------
        _auto_column_widths(sheet)

        # ---------- Style engine (per-sheet or global) ----------
        sheet_style = _parse_excel_style(sdef.get("style")) or global_style
        if sheet_style:
            full_data = [headers] + data_rows
            if sheet_style.column_widths:
                for col_letter, width in sheet_style.column_widths.items():
                    sheet.column_dimensions[col_letter].width = width
            _apply_style_engine(sheet, full_data, sheet_style, title_row=False)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =====================================================
#  EXC-04: extract_excel_named_range
# =====================================================

def _parse_cell_range(cell_range: str) -> tuple[int, int, int, int]:
    """
    解析 'A1:D10' 格式的单元格范围，返回 (start_row, end_row, start_col, end_col)
    """
    from openpyxl.utils import column_index_from_string

    match = re.match(r"^([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)$", cell_range)
    if not match:
        raise ValueError(f"无法解析单元格范围: {cell_range}")

    start_col = column_index_from_string(match.group(1))
    start_row = int(match.group(2))
    end_col = column_index_from_string(match.group(3))
    end_row = int(match.group(4))

    return start_row, end_row, start_col, end_col


def extract_excel_range(
    source_excel_url: str,
    sheet_name: str = "Sheet1",
    cell_range: Optional[str] = None,
    keyword: Optional[str] = None,
) -> dict:
    """
    从远程 Excel 中精准提取局部数据。

    Args:
        source_excel_url: 文件 URL
        sheet_name:       Sheet 名称
        cell_range:       单元格范围 (如 'A1:D10')
        keyword:          关键字搜索

    Returns:
        dict: {"sheet_name", "headers", "data", "total_rows"}
    """
    excel_data = _download_excel_from_url(source_excel_url)
    wb = openpyxl.load_workbook(excel_data, data_only=True)

    # 确定目标 Sheet
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
        sheet_name = sheet.title

    # ---------- 关键字搜索模式 ----------
    if keyword:
        return _extract_by_keyword(sheet, sheet_name, keyword)

    # ---------- 精确范围模式 ----------
    if cell_range:
        start_row, end_row, start_col, end_col = _parse_cell_range(cell_range)
    else:
        # 全量读取
        start_row = 1
        end_row = sheet.max_row
        start_col = 1
        end_col = sheet.max_column

    # 提取表头（第一行作为 headers）
    headers = []
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=start_row, column=col).value
        headers.append(str(val) if val is not None else "")

    # 提取数据行
    data = []
    for row in range(start_row + 1, end_row + 1):
        row_data = []
        for col in range(start_col, end_col + 1):
            val = sheet.cell(row=row, column=col).value
            row_data.append(val)
        data.append(row_data)

    return {
        "sheet_name": sheet_name,
        "headers": headers,
        "data": data,
        "total_rows": len(data),
    }


def _extract_by_keyword(sheet, sheet_name: str, keyword: str) -> dict:
    """
    基于关键字在 Sheet 中定位区域：
    找到包含关键字的单元格，以其所在行为起始行，向下读取连续有数据的行。
    """
    found_row = None
    found_col = None

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value and keyword in str(cell.value):
                found_row = cell.row
                found_col = cell.column
                break
        if found_row:
            break

    if found_row is None:
        return {
            "sheet_name": sheet_name,
            "headers": [],
            "data": [],
            "total_rows": 0,
        }

    # 从找到的位置开始，读取直到遇到完全空行
    start_col = 1
    end_col = sheet.max_column

    # 表头为 found_row
    headers = []
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=found_row, column=col).value
        headers.append(str(val) if val is not None else "")

    # 数据行
    data = []
    for row_num in range(found_row + 1, sheet.max_row + 1):
        row_data = []
        all_none = True
        for col in range(start_col, end_col + 1):
            val = sheet.cell(row=row_num, column=col).value
            row_data.append(val)
            if val is not None:
                all_none = False
        if all_none:
            break  # 遇到全空行就停止
        data.append(row_data)

    return {
        "sheet_name": sheet_name,
        "headers": headers,
        "data": data,
        "total_rows": len(data),
    }
